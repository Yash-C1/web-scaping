from bs4 import BeautifulSoup
from selenium import webdriver
import platform
import openpyxl
import os
import time
import requests
import random
import argparse
from discord import Webhook, RequestsWebhookAdapter
import traceback
from selenium.webdriver.common.by import By


class Scraper():
    def __init__(self):

        if platform.system()=='Windows':
            self.driver_path = "./chromedriver.exe"
        else:
            self.driver_path = "./chromedriver"

        self.chrome_options = webdriver.ChromeOptions()
        # PROXY = "127.0.0.1:24000"
        # self.chrome_options.add_argument('--proxy-server=%s' % PROXY)
        self.chrome_options.add_argument('--incognito')

        self.browser = self.get_webdriver()

        self.scraped_data = []

        self.existing_ids = []

        ## output sheet
        self.output_sheet = "pagesjaunes_output.xlsx"

        ## restarts webdriver when counter >= 100 and resets counter to 0
        self.counter = 0

        ## discord webhook
        self.progress_webhook = Webhook.from_url("https://discord.com/api/webhooks/883242626873172018/bI-5Nndc-uHvQqUOw0NreNLO2jwvNfCxcwRouIhSdC49OBtyrzW6Qx2ZhbYVr9oArr8N", adapter = RequestsWebhookAdapter())
        self.errors_webhook = Webhook.from_url("https://discord.com/api/webhooks/885079040585781249/nXYkngiZteIsSP4GtJqXg9zSfDHs5l-6FvfqZw-GFBr96vlnmYyYpEFlbwPrYHqMjNXr", adapter = RequestsWebhookAdapter())
        self.region = None
        
        ## handle exceptions
        self.captcha_encountered = False
        self.retry_counter = 0

        ##pagenos
        self.page_nos = None
        self.current_pageno = 1
        self.start_scraping_at_page_no = 2

    def captcha_solve(self):
        API_KEY = "14b952e64f255d96cfb78eef9feb549c"
        data_sitekey = '33f96e6a-38cd-421b-bb68-7806e1764460'
        page_url =self.browser.current_url
        time.sleep(10)
        u1 = f"https://2captcha.com/in.php?key={API_KEY}&method=hcaptcha&sitekey={data_sitekey}&pageurl={page_url}"
        r1 = requests.get(u1)
        try:
            print(r1.json())
            rid = r1.json().get("request")
        except:
            print(r1.text)
            rid = r1.text[3:]
        print("rid: ",rid)
        u2 = f"https://2captcha.com/res.php?key={API_KEY}&action=get&id={int(rid)}&json=1"
        time.sleep(5)
        count = 0
        while True:
            count = count+1
            r2 = requests.get(u2)
            print(r2.json())
            if r2.json().get("status") == 1:
                form_tokon = r2.json().get("request")
                break
            if count > 20:
                break
            time.sleep(5)
        wirte_tokon_js = f'document.getElementsByName("h-captcha-response")[0].value="{form_tokon}";'
        submit_js = 'document.getElementById("challenge-form").submit();'
        self.browser.execute_script(wirte_tokon_js)
        time.sleep(3)
        self.browser.execute_script(submit_js)
        time.sleep(5)
        self.errors_webhook.send(f"Captcha bypassed by captcha2 at --url: {page_url}")
        time.sleep(10)

    def get_webdriver(self):
        driver = webdriver.Chrome(executable_path=self.driver_path, options=self.chrome_options)

        return driver

    def extract_details(self, url):
        try:
            self.browser.get(url)

            time.sleep(3) ## wait to finish loading

            page_source = self.browser.page_source
            page_soup = BeautifulSoup(page_source, "html.parser")
        
        except Exception as e:
            traceback.print_exc()
            self.retry_counter += 1
            if self.retry_counter > 2:
                self.retry_counter = 0 ## reset counter
                raise e
            else:
                print("extract_details(): Retring..")
                return self.extract_details(url)

        self.retry_counter = 0 ## reset counter

        ## id
        id = url.split('/')[4]

        ## company_name
        try:
            nom_societe = page_soup.find('div', attrs={"class": "denom"}).h1.text
        except:
            if page_soup.find('input', attrs={"name": "cf_captcha_kind"}):
                print(f"Captcha page encountered.. Retrying! url: {url}") ## info
                self.captcha_encountered = True
                try:
                    self.errors_webhook.send(f"@here Captcha encountered while getting pros detail. Retrying in 5sec --region: {self.region}")
                except Exception as e:
                    print(e.__class__)
                    pass

                time.sleep(5) ##slese
                try:
                    self.captcha_solve()
                    try:
                        nom_societe = page_soup.find('div', attrs={"class": "denom"}).h1.text
                    except:
                        return self.extract_details(url)    
                except Exception as e:
                    print("solver failed: ",e)
                    return self.extract_details(url)
            else:
                time.sleep(4)
                return

        if self.captcha_encountered:
            self.captcha_encountered = False
            try:
                self.errors_webhook.send(f"Captcha bypassed. --region: {self.region}")
            except Exception as e:
                print(e.__class__)
                pass
        try:    
            ## address
            adr_txt = page_soup.find('div', attrs={'class': 'address-container marg-btm-s'}).find('a', attrs={'class': 'teaser-item black-icon address streetAddress clearfix map-click-zone pj-lb pj-link'}).text
            adr = adr_txt.split(',')
            adresse = adr[0].replace("Localisation", "")
            
            ## postal-code
            code_postal = adr[-1].strip().split(" ")[0]
            
            ## city
            ville = " ".join(adr[-1].strip().split(" ")[1:])
        except:
            adresse = ""
            code_postal = ""
            ville = ""
            
        ## type (heading)
        try:
            rubrique = self.browser.find_element(By.XPATH,"//div[@class = 'col-sm-6 col-md-7']//span[@class = 'activite']").text
        except:
            rubrique = ""

        ## siret, tva_intra, principaux_dirigeants
        zoneB2B = page_soup.find('div', attrs={'id': 'zoneB2B'})
        try:
            siret = zoneB2B.find('dl', attrs={'class': 'info-etablissement marg-btm-s zone-b2b txt_sm'}).find('dd').text.strip()
        except Exception:
            siret = ""
        
        tva_intra = ""
        try:
            for i, dt in enumerate(zoneB2B.find('dl', attrs={'class': 'info-entreprise marg-btm-s zone-b2b txt_sm'}).find_all('dt')):
                if dt.text == "TVA intracommunautaire":
                    tva_intra = zoneB2B.find('dl', attrs={'class': 'info-entreprise marg-btm-s zone-b2b txt_sm'}).find_all('dd')[i].text.strip()
                    break
        except Exception:
            pass

        principaux_dirigeants = ""
        try:
            for i, dt in enumerate(zoneB2B.find('dl', attrs={'class': 'info-entreprise marg-btm-s zone-b2b txt_sm'}).find_all('dt')):
                if dt.text == "Principaux dirigeants":
                    principaux_dirigeants = zoneB2B.find('dl', attrs={'class': 'info-entreprise marg-btm-s zone-b2b txt_sm'}).find_all('dd')[i].text.strip()
                    break
        except Exception:
            pass

        ## rating
        try:
            rating = self.browser.find_element(By.XPATH,"//div[3]/a/span[1]/span/strong").text
        except Exception:
            rating = ""
        ## review
        try:
            review = self.browser.find_element(By.XPATH,"//div[3]/a/span[2]/span").text.split(" ")[2]
        except Exception:
            review = ""
        # contacts-div
        try:
            contact_div = page_soup.find("div", attrs={"class": "fd-bloc bloc-coordonnees"})
        except:
            contact_div = None
        if contact_div:
            try:
                contact_numbers_spans = contact_div.find("div", attrs={"id": "coord-liste-numero_1"}).find_all("span", attrs={"class": "nb-phone"})
            except:
                pass
        try:
            for span in contact_numbers_spans:
                if span.find("span", attrs={"class": "num-tel-label"}).i.span.text == "Numéro de Téléphone":
                    telephone = span.find("span", attrs={"class": "coord-numero noTrad"}).text.strip()
                    break
                else:
                    telephone = ""
        except Exception:
            try:
                telephone = self.browser.find_element(By.XPATH,"//div[@class='fd-bloc bloc-coordonnees']//span[@class='nb-phone']//span[@class='arcep-hidden-phone']//div[@class='num num-arcep']").get_attribute("innerHTML")
            except:
                telephone = ""
        try:
            for span in contact_numbers_spans:
                if span.find("span", attrs={"class": "num-tel-label"}).i.span.text == "Numéro de Mobile":
                    mobile = span.find("span", attrs={"class": "coord-numero noTrad"}).text.strip()
                    break
                else:
                    mobile = ""
        except Exception:
            mobile = ""
        try:
            for span in contact_numbers_spans:
                if span.find("span", attrs={"class": "num-tel-label"}).span.text == "FAX":
                    fax = span.find("span", attrs={"class": "coord-numero noTrad"}).text.strip()
                    break
                else:
                    fax = ""
        except Exception:
            fax = ""

        ## website
        try:
            website = self.browser.find_element(By.XPATH,"//div[@class='bloc-info-sites-reseaux']//a[@class='SITE_EXTERNE pj-lb pj-link']/span[@class='value']").get_attribute("innerHTML")
        except Exception:
            website = ""

        ## activities
        try:
            activities = "|".join([a.text.strip() for a in page_soup.find("div", attrs={"class": "multi-activites"}).ul.find_all('a')])
        except Exception:
            activities = ""

        ## prestations 
        try:
            prestations = "|".join([span.text for span in page_soup.find("div", attrs={"class": "ligne prestations marg-btm-m generique"}).ul.find_all('span')])
        except Exception:
            prestations = ""

        ## description
        try:
            description = page_soup.find("div", attrs={"id": "teaser-description"}).find("div", attrs={"class": "description pf_description"}).p.text
        except:
            description = ""

        ## schedule
        try:
            horaires = "|".join([f"{li.p.text}: {li.ul.li.text.strip()}" for li in page_soup.find("div", attrs={"id": "infos-horaires"}).ul.find_all("li", attrs={"class": "horaire-ouvert"})])
        except:
            try:
                horaires = page_soup.find("div", attrs={"id": "infos-horaires"}).p.text
            except:
                horaires = ""
        
        ## 
        try:
            budget = "|".join([li.text.strip() for li in page_soup.find("div", attrs={"id": "tarif-generique"}).ul.find_all("li")])
        except:
            budget = ""
        ##
        try:
            tarif_nuit = page_soup.find("div", attrs={"id": "tarif-hotel"}).p.span.text
        except:
            tarif_nuit = ""
        
        ## 
        try:
            cuisine = "|".join([li.text for li in page_soup.find("div", attrs={"class": "bloc-info-cuisine"}).ul.find_all("li")])
        except:
            cuisine = ""
        
        ##
        try:
            ambiance = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "bloc-info-ambiance"}).ul.find_all("li")])
        except:
            ambiance = ""

        ##
        try:
            references_et_guides = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "zone-info-guides"}).div.ul.find_all("li")])
        except:
            references_et_guides = ""

        ##
        try:
            moyens_de_paiement = "|".join([img['alt'] for img in page_soup.find("div", attrs={"class": "zone-info-moyen-paiement"}).div.ul.find_all("img")])
        except:
            moyens_de_paiement = ""
        
        ## stars
        try:
            nbre_etoile_hotel = page_soup.find("div", attrs={"class": "bloc-info-categories col-sm-6 col-lg-4"}).ul.li.find("span", attrs={"class": "categorie-libelle"}).text.split(" ")
            if len(nbre_etoile_hotel) == 3:
                nbre_etoile_hotel = nbre_etoile_hotel[1]
            elif len(nbre_etoile_hotel) == 2:
                nbre_etoile_hotel = nbre_etoile_hotel[0]
        except:
            nbre_etoile_hotel = ""
        
        ## 
        try:
            nbre_chambres = page_soup.find("div", attrs={"id": "zone-info-nb-chambres"}).div.span.text
        except:
            nbre_chambres = ""
        ##
        try:
            marque = "|".join([li.text for li in page_soup.find("div", attrs={"class": "ligne marques marg-btm-m"}).ul.find_all("li")])
        except:
            marque = ""
        
        ##
        try:
            capacite_accueil = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "zone-info-capacites"}).div.ul.find_all("li")])
        except:
            capacite_accueil = ""

        ##
        try:
            info_pratique = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "zone-info-prestations-pratiques"}).div.ul.find_all("li")])
        except:
            info_pratique = ""

        ##
        try:
            services_loisirs = page_soup.find("div", attrs={"class": "zone-info-services-loisirs"}).p.text
        except:
            services_loisirs = ""
        
        try:
            clientele = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "zone-info-clientele"}).div.ul.find_all("li")])
        except:
            clientele = ""

        try:
            nom_du_chef = page_soup.find("div", attrs={"class": "zone-info-chef"}).span.text
        except:
            nom_du_chef = ""
        
        try:
            produits = "|".join([li.span.text for li in page_soup.find("div", attrs={"class": "ligne produits"}).ul.find_all("li")])
        except:
            produits = ""
        
        try:
            dernire_modif = page_soup.find("div", attrs={"class": "maj-date txt_xs"}).text
        except:
            dernire_modif = ""

        ## remaining
        formules = ""

        row = [url, id, nom_societe, adresse, code_postal, ville, rubrique, siret, dernire_modif, tva_intra, principaux_dirigeants, rating, review, telephone, mobile, fax, website, activities, prestations, produits, description, horaires, budget, tarif_nuit, cuisine, ambiance, formules, nom_du_chef, references_et_guides, moyens_de_paiement, nbre_etoile_hotel, nbre_chambres, marque, capacite_accueil, info_pratique, services_loisirs, clientele]

        self.scraped_data.append(row)

    def init_sheets(self):

        if not os.path.exists(self.output_sheet):
            ## Column Headers
            wb = openpyxl.Workbook()
            sheet = wb.active
            header = ["url", "id", "nom_societe", "adresse", "code_postal", "ville", "rubrique", "siret", "dernire_modif", "tva_intra", "principaux_dirigeants", "rating", 'review', "telephone", 'mobile', "fax", "website", "activities", 'prestations', "produits", "description", "horaires", "budget", "tarif_nuit", "cuisine", "ambiance", "formules", "nom_du_chef", "references_et_guides", "moyens_de_paiement", "nbre_etoile_hotel", "nbre_chambres", "marque", "capacite_accueil", "info_pratique", "services_loisirs", "clientele"]
            sheet.append(header)
            wb.save(filename=self.output_sheet)
            wb.close()
        else:
            wb = openpyxl.load_workbook(filename=self.output_sheet)
            ws = wb.active

            for col in ws['B']:
                self.existing_ids.append(col.value) ## read existing values
            
            wb.close()


    def write_to_sheet(self):
        print("writing to sheet") ## info

        wb = openpyxl.load_workbook(self.output_sheet)
        ws = wb.active
        for row in self.scraped_data:
            ws.append(row)
        
        wb.save(filename=self.output_sheet)
        wb.close()

        self.scraped_data.clear()

        print("writing done") ## info

    def get_professionals_data(self, city_link):
        print(city_link)
        try:
            time.sleep(3)
            self.browser.get(city_link)
            time.sleep(3)
            print("currentUrl",self.browser.current_url)
            if city_link != self.browser.current_url:
                print("page doesn't exist, returning")
                time.sleep(random.randint(4,6))
                return
            page_source = self.browser.page_source
            page_soup = BeautifulSoup(page_source, "html.parser")
            if self.current_pageno == 1:
                print("entered")
                try:
                    pagination_elem = page_soup.find('p', attrs={'class': 'pagination col-xs-12 text-center'})
                    all_pages = pagination_elem.find_all('a', attrs={'class': 'pj-link'})
                except:
                    print("No data Found !")
                    return
                if not all_pages:
                    print("left")
                    return
                for pages in all_pages:
                    if int(pages['title']) > self.current_pageno:
                        print('started at page : ',pages['title'])
                        self.current_pageno = int(pages['title'])

                        ak_lt = city_link.split('/')
                        if ak_lt[-1] ==  "professionnels":
                            city_link += f"/{self.current_pageno}"

                        return self.get_professionals_data(city_link)
        except Exception as e:
            traceback.print_exc()
            self.retry_counter += 1
            if self.retry_counter > 2:
                self.retry_counter = 0 ## reset counter
                raise e
            else:
                print("get_professionals_data(): Retring..")
                return self.get_professionals_data(city_link)
        
        self.retry_counter = 0 ## reset counter

        try:
            a_tags = page_soup.find('ul', attrs={'class': 'col-xs-12 liste2colonnes'}).find_all('a', attrs={'class': 'pj-link'})
        except:
            if page_soup.find('input', attrs={"name": "cf_captcha_kind"}):
                print("Captcha page encountered.. Retrying!") ## info
                self.captcha_encountered = True
                try:
                    self.errors_webhook.send(f"@here Captcha encountered while getting professional links. Retrying in 5sec --region: {self.region}")
                except Exception as e:
                    print(e.__class__)
                    pass
                time.sleep(5) ##slese
                try:
                    self.captcha_solve()
                    try:
                        a_tags = page_soup.find('ul', attrs={'class': 'col-xs-12 liste2colonnes'}).find_all('a', attrs={'class': 'pj-link'})
                        print("captcha solved",a_tags)
                    except:
                        return self.get_professionals_data(city_link)    
                except Exception as e:
                    print("solver not working: ",e)
                    return self.get_professionals_data(city_link)
            else:
                print("No data found..!") ## info
                time.sleep(random.randint(4,6)) ##sleep
                return
        if self.captcha_encountered:
            self.captcha_encountered = False
            try:
                self.errors_webhook.send(f"Captcha bypassed. --region: {self.region}")
            except Exception as e:
                print(e.__class__)
                pass
        
        urls = [f"https://www.pagesjaunes.fr{a['href']}" for a in a_tags]
        
        for i, url in enumerate(urls):

            id = url.split('/')[4]

            if id in self.existing_ids:
                continue

            print(f"({i + 1}/{len(urls)}): {url}") ## info 

            time.sleep(random.randint(4,5)) ## sleep
                
            self.extract_details(url)
            ## update counter
            self.counter += 1

            if len(self.scraped_data) == 50:
                self.write_to_sheet()

            if self.counter >= 100:
                ## close existing webdriver and create a new one
                print("--restarting chromedriver--")
                self.browser.close()
                self.browser = self.get_webdriver()
                self.counter = 0

        if not self.page_nos:
            print("entered")
            pagination_elem = page_soup.find('p', attrs={'class': 'pagination col-xs-12 text-center'})
            all_pages = pagination_elem.find_all('a', attrs={'class': 'pj-link'})


            if not all_pages:
                print("left")
                return

            print('all : ',all_pages)
            for pages in all_pages:
                print('title : ',pages['title'])
                if int(pages['title']) > self.current_pageno:
                    print('title : ',pages['title'])
                    self.current_pageno = int(pages['title'])

                    ak_lt = city_link.split('/')
                    if ak_lt[-1] !=  "professionnels":
                        city_link = city_link[:-len(ak_lt[-1])]
                    city_link += f"{self.current_pageno}"

                    return self.get_professionals_data(city_link)


    
    def get_city_by_letter(self, region_link, city_starts_with):
        try:
            self.browser.get(region_link)

            page_source = self.browser.page_source
            page_soup = BeautifulSoup(page_source, "html.parser")
        except Exception as e:
            traceback.print_exc()
            self.retry_counter += 1
            if self.retry_counter > 2:
                self.retry_counter = 0 ## reset counter
                raise e
            else:
                print("get_city_by_letter(): Retring..")
                return self.get_city_by_letter(region_link, city_starts_with)
        
        self.retry_counter = 0 ## reset counter

        try:
            h3 = page_soup.find_all("h3", attrs={"class": "col-xs-12"})
            p_tags = page_soup.find("div", attrs={"class": "row"}).find_all("p", attrs={"class": "col-xs-12"})
        except Exception as e:
            if page_soup.find('input', attrs={"name": "cf_captcha_kind"}):
                print("Captcha page encountered while getting cities.. Retrying!") ## info
                self.captcha_encountered = True
                try:
                    self.errors_webhook.send(f"@here Captcha encountered while getting cities. Retrying in 5sec --region: {self.region}")
                except Exception as e:
                    print(e.__class__)
                    pass
                time.sleep(5) ##slese
                try:
                    self.captcha_solve()
                    try:
                        h3 = page_soup.find_all("h3", attrs={"class": "col-xs-12"})
                        p_tags = page_soup.find("div", attrs={"class": "row"}).find_all("p", attrs={"class": "col-xs-12"})
                    except:
                        return self.get_city_by_letter(region_link, city_starts_with)

                except:
                    return self.get_city_by_letter(region_link, city_starts_with)
            else:
                raise e

        if self.captcha_encountered:
            self.captcha_encountered = False
            try:
                self.errors_webhook.send(f"Captcha bypassed. --region: {self.region}")
            except Exception as e:
                print(e.__class__)
                pass

        if not h3:
            print("No city by letter Links found")
            time.sleep(2)

        ## find links corresponding to each letters
        for i, h in enumerate(h3):
            try:
                ## if city_starts_with is specified, skip every other letters
                current_letter = h.a.text.split("(voir toutes les villes)")[0].strip().lower()
                if city_starts_with:
                    if current_letter not in city_starts_with.lower():
                        continue
            except:
                current_letter = h.text.strip().lower()
                if city_starts_with:
                    if current_letter not in city_starts_with.lower():
                        continue

            if h.find("a"):
                cities_by_letter = f"https://www.pagesjaunes.fr{h.find('a')['href']}"

                time.sleep(3) ##sleep

                while True:
                    self.browser.get(cities_by_letter)
                
                    page_source = self.browser.page_source
                    page_soup = BeautifulSoup(page_source, "html.parser")

                    if not page_soup.find('input', attrs={"name": "cf_captcha_kind"}):
                        break
                    else:
                        self.captcha_solve()
                    
                    time.sleep(3)

                ul_lists = page_soup.find_all('ul', attrs={"class": "liste3colonnes col-xs-12 col-sm-4 marg-btm-l"})

                ## get links to page containing professionnels links
                city_professionnels_links = []
                for ul in ul_lists:
                    rows = ul.find_all('li', attrs={"class": "marg-btm-xxs"})
                    for row in rows:
                        if current_letter:
                            if row.find('a', attrs={"class": "pj-link"}).text[0].lower() not in current_letter:
                                continue 
                        city_link = row.find('a', attrs={"class": "pj-link"})['href']
                        city_professionnels_links.append(f"https://www.pagesjaunes.fr{city_link}/professionnels")
                
                # print(city_professionnels_links)
                for i, city_link in enumerate(city_professionnels_links):
                    print(f"City({i + 1}/{len(city_professionnels_links)}): {city_link}")
                    if self.page_nos:
                        for pg_no in self.page_nos:
                            self.get_professionals_data(city_link + f"/{pg_no}")
                    else:
                        self.get_professionals_data(city_link)

            else:
                p = p_tags[i + 1]
                city_professionnels_links = [f"https://www.pagesjaunes.fr{a['href']}/professionnels/{self.start_scraping_at_page_no}" for a in p.find_all("a")]
                # print(city_professionnels_links)
                for i, city_link in enumerate(city_professionnels_links):
                    print(f"City({i + 1}/{len(city_professionnels_links)}): {city_link}")
                    if self.page_nos:
                        for pg_no in self.page_nos:
                            self.get_professionals_data(city_link + f"/{pg_no}")
                    else:
                        self.get_professionals_data(city_link)
            
        ## write any unwritten data
        if self.scraped_data:
            self.write_to_sheet()
    
    def get_region(self, scrape_region=None, city_starts_with=None):
        base_link = "https://www.pagesjaunes.fr/"

        self.init_sheets() ##

        try:
            self.browser.get(base_link)

            page_source = self.browser.page_source
            page_soup = BeautifulSoup(page_source, "html.parser")
        except Exception as e:
            traceback.print_exc()
            self.retry_counter += 1
            if self.retry_counter > 2:
                self.retry_counter = 0 ## reset counter
                raise e
            else:
                print("get_region(): Retring..")
                return self.get_region(scrape_region=scrape_region, city_starts_with=city_starts_with)
        
        self.retry_counter = 0 ## reset counter

        try:
            region_list = page_soup.find("div", attrs={"class": "region marg-btm-xl row"}).ul
        except Exception as e:
            if page_soup.find('input', attrs={"name": "cf_captcha_kind"}):
                print("Captcha page encountered while getting region.. Retrying!") ## info
                self.captcha_encountered = True
                try:
                    self.errors_webhook.send(f"@here Captcha encountered while getting region. Retrying in 5sec --region: {self.region}")
                except Exception as e:
                    print(e.__class__)
                    pass
                time.sleep(5) ##sleep
                try:
                    self.captcha_solve()
                    try:
                        region_list = page_soup.find("div", attrs={"class": "region marg-btm-xl row"}).ul
                    except:
                        return self.get_region(scrape_region, city_starts_with)

                except:
                    return self.get_region(scrape_region, city_starts_with)
            else:
                raise e

        if self.captcha_encountered:
            self.captcha_encountered = False
            try:
                self.errors_webhook.send(f"Captcha bypassed. --region: {self.region}")
            except Exception as e:
                print(e.__class__)
                pass

        for li in region_list.find_all("li", attrs={"class": "clearfix"}):
            region = li.a.text
            self.region = region.lower()
            if scrape_region:
                if scrape_region.lower() != region.lower():
                    continue

            region_link = f"https://www.pagesjaunes.fr{li.a['href']}"

            self.get_city_by_letter(region_link, city_starts_with)


    def start(self, args):

        timestamp = time.strftime("%H-%M-%Y-%m")

        page_nos = args.page_nos
        if page_nos:
            if ',' in page_nos:
                page_nos = page_nos.split(',')
            else:
                page_nos = [page_nos]

        self.page_nos = page_nos

        if args.professional_url:
            self.output_sheet = f"output-{timestamp}.xlsx"

            self.init_sheets()

            ## scrape only given url
            if self.page_nos:
                for pg_no in self.page_nos:
                    self.get_professionals_data(args.professional_url + f"/{pg_no}")
            else:
                self.extract_details(args.professional_url)
            
                self.write_to_sheet()

            self.browser.close()
            return "pass"
        
        if args.region:
            region = args.region.replace(' ', '-')
            self.region = region
            if args.city_starts_with:
                starts_with = args.city_starts_with.replace(' ', '').replace(',', '-')

                self.output_sheet = f"{region}_city_{starts_with}.xlsx"
                ## scrape only cities starting with given letter under given region
                self.get_region(args.region, args.city_starts_with)
            else:
                self.output_sheet = f"region_{region}.xlsx"
                ## scrape only given region
                self.get_region(scrape_region=args.region)
            
            self.browser.close()
            return "pass"
        
        if args.city_slug:
            self.output_sheet = f"{args.city_slug}_output-{timestamp}.xlsx"

            self.init_sheets()
            ## get link using given city slug
            city_professionnels_link = f"https://www.pagesjaunes.fr/annuaire/{args.city_slug}/professionnels/{self.start_scraping_at_page_no}"
            ## scrape only given city
            self.get_professionals_data(city_professionnels_link)

            ## write any unwritten data
            if self.scraped_data:
                self.write_to_sheet()

            self.browser.close()
            return "pass"
        
        ## if no cli args
        self.get_region()

        self.browser.close()
        return "pass"


if __name__ == "__main__":
    ## handle cli arguments
    argparser = argparse.ArgumentParser()

    argparser.add_argument('--region', dest='region', type=str, help='region')
    argparser.add_argument('--cityStartsWith', dest='city_starts_with', type=str, help='city starts with letter _')
    argparser.add_argument('--citySlug', dest='city_slug', type=str, help='city slug')
    argparser.add_argument('--professionalUrl', dest='professional_url', type=str, help='professional url to scrape')
    argparser.add_argument('--pageNos', dest='page_nos', type=str, help='page nos to scrape')

    args = argparser.parse_args()

    scraper = Scraper()

    # scraper.start(args)

    ## send start notification to discord
    # scraper.progress_webhook.send(f"""[ ] Started Scrape
    # cli-args passed:
    # {'':2}--region: {args.region}
    # {'':2}--cityStartsWith: {args.city_starts_with}
    # {'':2}--citySlug: {args.city_slug}
    # {'':2}--professionalUrl: {args.professional_url}
    # {'':2}--pageNos: {args.page_nos}""")

    status = None
    while status != "pass":
        try:
            ## start scraping
            status = scraper.start(args)
        except Exception as e:
            traceback.print_exc()
            try:
                scraper.errors_webhook.send(f"""@everyone [#] Scraper encountered error. retrying in 20s..
        cli-args passed:
        {'':2}--region: {args.region}
        {'':2}--cityStartsWith: {args.city_starts_with}
        {'':2}--citySlug: {args.city_slug}
        {'':2}--professionalUrl: {args.professional_url}
        {'':2}--pageNos: {args.page_nos}""")
                scraper.errors_webhook.send(e)
            except Exception as e:
                print(e.__class__)
                pass

        try:
            scraper.browser.close()
        except:
            pass

        scraper.browser = scraper.get_webdriver()
        time.sleep(20)

    ## send completion notification to discord
    # scraper.progress_webhook.send(f"""@everyone [#] Scrape Complete
    # cli-args passed:
    # {'':2}--region: {args.region}
    # {'':2}--cityStartsWith: {args.city_starts_with}
    # {'':2}--citySlug: {args.city_slug}
    # {'':2}--professionalUrl: {args.professional_url}
    # {'':2}--pageNos: {args.page_nos}""")
    